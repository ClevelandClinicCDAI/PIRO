"""FastAPI routes for the PIRO Extraction Suite.

All endpoints require authentication. Provider selection is server-side only.
"""

from __future__ import annotations

import csv
import io
import json
import asyncio
from typing import Annotated, Any, Dict, List, Optional

from core.auth_bearer import JWTBearer
from core.constants import Constants
from core.llm_client import FieldExtraction, get_llm_client
from core.security_user import get_current_user_id, get_current_user_nuid, get_current_user_role
from db.repository.extraction import (
    add_cases_to_queue,
    bulk_approve_high_confidence,
    build_labelled_report_text,
    create_run,
    create_session,
    delete_session,
    get_case_text_for_extraction,
    get_case_text_segments,
    get_extraction_status,
    get_incorrect_case_ids,
    get_latest_run,
    get_low_confidence_case_ids,
    get_queue,
    get_result_by_id,
    get_results_for_session,
    get_run,
    get_session,
    list_user_sessions,
    remove_from_queue,
    reset_queue_statuses,
    update_queue_item_status,
    update_result_review,
    update_run_status,
    update_session_name,
    update_session_schema,
    update_session_status,
    upsert_result,
)
from db.repository.cohort import get_cohort
from db.repository.search import get_search
from urllib.parse import parse_qs, urlparse
import json as _json
from db.session import SessionLocal, get_db, get_solr
from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query
from fastapi.responses import JSONResponse, StreamingResponse
from logger import logger
from sqlalchemy.orm import Session
from viewmodel.extraction import (
    CaseTextVM,
    CaseCommentSegmentVM,
    ExtractionPreviewRequest,
    ExtractionPreviewVM,
    ExtractionPreviewFieldVM,
    ExtractionQueueAdd,
    ExtractionQueueFromSearch,
    ExtractionQueueItemVM,
    ExtractionResultPatch,
    ExtractionResultVM,
    ExtractionRunRequest,
    ExtractionRunVM,
    ExtractionSessionCreate,
    ExtractionSessionUpdate,
    ExtractionSessionVM,
    ExtractionStatusVM,
    FieldSuggestionRequest,
    FieldSuggestionVM,
)

router = APIRouter()

_ALLOWED_ROLES = [
    Constants.RoleAdmin,
    Constants.RoleDemoAdmin,
    Constants.RoleAnalyst,
    Constants.RoleUser,
]


def _require_session_ownership(session_id: int, user_id: int, db: Session):
    """Raise 404 if session doesn't exist, 403 if it belongs to a different user."""
    sess = get_session(session_id, db)
    if sess is None:
        raise HTTPException(status_code=404, detail="Extraction session not found")
    if sess.UserId != user_id:
        raise HTTPException(status_code=403, detail="Access denied")
    return sess


# ──────────────────────────────────────────────────────────────────────────────
# Session endpoints
# ──────────────────────────────────────────────────────────────────────────────

@router.post(
    "/session",
    dependencies=[Depends(JWTBearer(_ALLOWED_ROLES))],
    response_model=ExtractionSessionVM,
)
async def create_extraction_session(
    payload: ExtractionSessionCreate,
    current_user: Annotated[str, Depends(get_current_user_nuid)],
    current_user_id: Annotated[int, Depends(get_current_user_id)],
    db: Session = Depends(get_db),
):
    sess = create_session(
        name=payload.name,
        user_id=current_user_id,
        user=current_user,
        db=db,
    )
    if payload.schema_definition:
        update_session_schema(sess.ExtractionSessionId, payload.schema_definition, current_user, db)
        db.refresh(sess)
    return sess


@router.get(
    "/sessions",
    dependencies=[Depends(JWTBearer(_ALLOWED_ROLES))],
    response_model=List[ExtractionSessionVM],
)
async def list_sessions(
    current_user_id: Annotated[int, Depends(get_current_user_id)],
    db: Session = Depends(get_db),
):
    return list_user_sessions(user_id=current_user_id, db=db)


@router.get(
    "/session/{session_id}",
    dependencies=[Depends(JWTBearer(_ALLOWED_ROLES))],
    response_model=ExtractionSessionVM,
)
async def get_extraction_session(
    session_id: int,
    current_user_id: Annotated[int, Depends(get_current_user_id)],
    db: Session = Depends(get_db),
):
    return _require_session_ownership(session_id, current_user_id, db)


@router.put(
    "/schema/{session_id}",
    dependencies=[Depends(JWTBearer(_ALLOWED_ROLES))],
    response_model=ExtractionSessionVM,
)
async def save_schema(
    session_id: int,
    payload: ExtractionSessionUpdate,
    current_user: Annotated[str, Depends(get_current_user_nuid)],
    current_user_id: Annotated[int, Depends(get_current_user_id)],
    db: Session = Depends(get_db),
):
    _require_session_ownership(session_id, current_user_id, db)
    if payload.name is not None:
        update_session_name(session_id, payload.name, current_user, db)
    if payload.schema_definition is not None:
        update_session_schema(session_id, payload.schema_definition, current_user, db)
    return get_session(session_id, db)


@router.delete(
    "/session/{session_id}",
    dependencies=[Depends(JWTBearer(_ALLOWED_ROLES))],
)
async def archive_session(
    session_id: int,
    current_user: Annotated[str, Depends(get_current_user_nuid)],
    current_user_id: Annotated[int, Depends(get_current_user_id)],
    db: Session = Depends(get_db),
):
    _require_session_ownership(session_id, current_user_id, db)
    delete_session(session_id, current_user, db)
    return {"detail": "Session archived"}


# ──────────────────────────────────────────────────────────────────────────────
# Queue endpoints
# ──────────────────────────────────────────────────────────────────────────────

@router.post(
    "/queue",
    dependencies=[Depends(JWTBearer(_ALLOWED_ROLES))],
    response_model=List[ExtractionQueueItemVM],
)
async def add_to_queue(
    payload: ExtractionQueueAdd,
    current_user: Annotated[str, Depends(get_current_user_nuid)],
    current_user_id: Annotated[int, Depends(get_current_user_id)],
    db: Session = Depends(get_db),
):
    _require_session_ownership(payload.session_id, current_user_id, db)
    added = add_cases_to_queue(
        session_id=payload.session_id,
        case_ids=payload.case_ids,
        user=current_user,
        db=db,
    )
    return get_queue(payload.session_id, db)


@router.post(
    "/queue/from-saved-search",
    dependencies=[Depends(JWTBearer(_ALLOWED_ROLES))],
    response_model=List[ExtractionQueueItemVM],
)
async def add_saved_search_to_queue(
    payload: ExtractionQueueFromSearch,
    current_user: Annotated[str, Depends(get_current_user_nuid)],
    current_user_id: Annotated[int, Depends(get_current_user_id)],
    db: Session = Depends(get_db),
    solr=Depends(get_solr),
):
    from core.search_util import filter_str_object
    from solr.repository.piro import search_Q
    from viewmodel.solr.search import SearchFilterVM

    _require_session_ownership(payload.session_id, current_user_id, db)

    saved = get_search(payload.search_id, db)
    if saved is None:
        raise HTTPException(status_code=404, detail="Saved search not found")

    # Parse the stored URL query string to reconstruct the filter array
    parsed_q = parse_qs(urlparse(saved.SearchQuery).query)
    raw_filters = _json.loads(parsed_q.get("searchFilter", ["[]"])[0])
    filters = [
        SearchFilterVM(
            field=f["field"],
            search=f["search"],
            category=f["category"],
            andcondition=f["andcondition"],
            displaysingular=f.get("displaysingular", ""),
        )
        for f in raw_filters
    ]

    adv_filter = filter_str_object(saved.AdvancedQuery or "{}")
    mrn = saved.MRN or ""

    docs = search_Q(
        input_arr=filters,
        input_adv=adv_filter,
        mrn=mrn,
        sortBy="accessiondate",
        sortOrder="desc",
        page=0,
        count=999999,
        db=db,
        solr=solr,
        finalRtf=False,
        fields="id,caseid",
    )

    case_ids = docs.get("caseIds", [])
    if not case_ids:
        total = docs.get("total", 0)
        detail = "Saved search returned no cases" if total == 0 else f"Search matched {total} documents but none had a caseId field"
        raise HTTPException(status_code=400, detail=detail)

    add_cases_to_queue(
        session_id=payload.session_id,
        case_ids=case_ids,
        user=current_user,
        db=db,
    )
    return get_queue(payload.session_id, db)


@router.get(
    "/queue/{session_id}",
    dependencies=[Depends(JWTBearer(_ALLOWED_ROLES))],
    response_model=List[ExtractionQueueItemVM],
)
async def get_queue_items(
    session_id: int,
    current_user_id: Annotated[int, Depends(get_current_user_id)],
    db: Session = Depends(get_db),
):
    _require_session_ownership(session_id, current_user_id, db)
    return get_queue(session_id, db)


@router.delete(
    "/queue/{session_id}/{case_id}",
    dependencies=[Depends(JWTBearer(_ALLOWED_ROLES))],
)
async def remove_queue_item(
    session_id: int,
    case_id: int,
    current_user_id: Annotated[int, Depends(get_current_user_id)],
    db: Session = Depends(get_db),
):
    _require_session_ownership(session_id, current_user_id, db)
    if not remove_from_queue(session_id, case_id, db):
        raise HTTPException(status_code=404, detail="Queue item not found")
    return {"detail": "Removed"}


# ──────────────────────────────────────────────────────────────────────────────
# Extraction run
# ──────────────────────────────────────────────────────────────────────────────

async def _run_extraction_job(
    session_id: int,
    run_id: int,
    user: str,
    role: str,
    case_ids: Optional[List[int]] = None,
) -> None:
    """Background task: extract cases for a session.

    If ``case_ids`` is given (validation run), only those cases are processed;
    all others remain at their current queue status.
    If ``case_ids`` is None (full run), all queue items are (re-)processed.
    """
    db = SessionLocal()
    try:
        update_run_status(run_id, "running", db)
        update_session_status(session_id, "running", db)

        llm = get_llm_client()
        from core.config import settings

        queue = get_queue(session_id, db)
        run = get_run(run_id, db)
        schema = json.loads(run.SchemaJson)

        # Determine which cases to process
        validation_set = set(case_ids) if case_ids is not None else None

        # For full runs, reset all queue items so previously completed validation
        # cases are reprocessed to produce a complete result set
        if validation_set is None:
            reset_queue_statuses(session_id, db)

        for queue_item in queue:
            if validation_set is not None and queue_item.CaseId not in validation_set:
                continue  # validation run: skip non-sampled cases
            if validation_set is None and queue_item.Status == "completed":
                continue  # full run: skip already done (shouldn't exist after reset)

            update_queue_item_status(queue_item.ExtractionQueueId, "running", db)
            try:
                labelled_text, segments = get_case_text_for_extraction(
                    queue_item.CaseId, db, role=role
                )

                if not labelled_text.strip():
                    update_queue_item_status(
                        queue_item.ExtractionQueueId, "failed", db,
                        error="No report text found for this case"
                    )
                    continue

                extraction = await llm.extract(labelled_text, schema)

                # Build a character-offset map for provenance anchoring
                offset_map: Dict[int, tuple] = {}
                running_offset = 0
                for seg in segments:
                    label_prefix = f"{seg.CommentType}:\n"
                    content_start = running_offset + len(label_prefix)
                    content_end = content_start + len(seg.CommentText)
                    offset_map[seg.Id] = (content_start, content_end, seg.CaseCommentId if hasattr(seg, 'CaseCommentId') else None)
                    running_offset = content_end + 2  # +2 for "\n\n"

                for field_name, fe in extraction.items():
                    # Find provenance location
                    source_comment_id = None
                    prov_start = None
                    prov_end = None

                    if fe.provenance and segments:
                        for seg in segments:
                            idx = seg.CommentText.find(fe.provenance)
                            if idx != -1:
                                source_comment_id = getattr(seg, 'CaseCommentId', None)
                                if seg.Id in offset_map:
                                    base = offset_map[seg.Id][0]
                                    prov_start = base + idx
                                    prov_end = prov_start + len(fe.provenance)
                                break

                    extracted_json = json.dumps(fe.value) if fe.value is not None else None

                    upsert_result(
                        run_id=run_id,
                        session_id=session_id,
                        case_id=queue_item.CaseId,
                        field_name=field_name,
                        extracted_value=extracted_json,
                        confidence=fe.confidence,
                        provenance_text=fe.provenance,
                        source_comment_id=source_comment_id,
                        provenance_start=prov_start,
                        provenance_end=prov_end,
                        user=user,
                        db=db,
                    )

                update_queue_item_status(queue_item.ExtractionQueueId, "completed", db)

            except Exception as e:
                logger.error(
                    f"Extraction failed for case {queue_item.CaseId}: {e}", exc_info=True
                )
                update_queue_item_status(
                    queue_item.ExtractionQueueId,
                    "failed",
                    db,
                    error=str(e)[:1000],
                )

        # Update run/session status
        status = get_extraction_status(session_id, db)
        final_status = "completed" if status["failed"] == 0 else "completed_with_errors"
        update_run_status(run_id, final_status, db)
        update_session_status(session_id, final_status, db)

    except Exception as e:
        logger.error(f"Extraction job failed for session {session_id}: {e}", exc_info=True)
        update_run_status(run_id, "failed", db, error=str(e)[:1000])
        update_session_status(session_id, "failed", db)
    finally:
        db.close()


@router.post(
    "/run",
    dependencies=[Depends(JWTBearer(_ALLOWED_ROLES))],
    response_model=ExtractionRunVM,
)
async def start_extraction(
    payload: ExtractionRunRequest,
    background_tasks: BackgroundTasks,
    current_user: Annotated[str, Depends(get_current_user_nuid)],
    current_user_id: Annotated[int, Depends(get_current_user_id)],
    current_role: Annotated[str, Depends(get_current_user_role)],
    db: Session = Depends(get_db),
):
    sess = _require_session_ownership(payload.session_id, current_user_id, db)

    if not sess.SchemaJson:
        raise HTTPException(status_code=400, detail="Schema is not defined for this session")

    queue = get_queue(payload.session_id, db)
    if not queue:
        raise HTTPException(status_code=400, detail="Queue is empty")

    # Concurrency guard: reject if a run is already active
    latest = get_latest_run(payload.session_id, db)
    if latest and latest.Status in ("pending", "running"):
        raise HTTPException(
            status_code=409,
            detail=f"A run is already {latest.Status}. Wait for it to finish before starting another.",
        )

    from core.config import settings

    llm_provider = settings.LLM_PROVIDER or "ollama"
    llm_model = settings.LLM_MODEL or "llama3.2"

    # For validation runs, randomly sample the requested number of cases
    sampled_case_ids: Optional[List[int]] = None
    actual_validation_size: Optional[int] = None
    if payload.run_type == "validation":
        import random as _random
        all_ids = [q.CaseId for q in queue]
        n = min(payload.validation_size, len(all_ids))
        sampled_case_ids = _random.sample(all_ids, n)
        actual_validation_size = n

    run = create_run(
        session_id=payload.session_id,
        schema_json=sess.SchemaJson,
        llm_provider=llm_provider,
        llm_model=llm_model,
        user=current_user,
        db=db,
        run_type=payload.run_type,
        validation_size=actual_validation_size,
    )

    background_tasks.add_task(
        _run_extraction_job,
        session_id=payload.session_id,
        run_id=run.ExtractionRunId,
        user=current_user,
        role=current_role,
        case_ids=sampled_case_ids,
    )

    return run


@router.get(
    "/status/{session_id}",
    dependencies=[Depends(JWTBearer(_ALLOWED_ROLES))],
    response_model=ExtractionStatusVM,
)
async def get_status(
    session_id: int,
    current_user_id: Annotated[int, Depends(get_current_user_id)],
    db: Session = Depends(get_db),
):
    _require_session_ownership(session_id, current_user_id, db)
    status = get_extraction_status(session_id, db)
    return status


# ──────────────────────────────────────────────────────────────────────────────
# Results endpoints
# ──────────────────────────────────────────────────────────────────────────────

@router.get(
    "/results/{session_id}",
    dependencies=[Depends(JWTBearer(_ALLOWED_ROLES))],
    response_model=List[ExtractionResultVM],
)
async def get_results(
    session_id: int,
    current_user_id: Annotated[int, Depends(get_current_user_id)],
    db: Session = Depends(get_db),
):
    _require_session_ownership(session_id, current_user_id, db)
    return get_results_for_session(session_id, db)


@router.patch(
    "/results/{result_id}",
    dependencies=[Depends(JWTBearer(_ALLOWED_ROLES))],
    response_model=ExtractionResultVM,
)
async def patch_result(
    result_id: int,
    payload: ExtractionResultPatch,
    current_user: Annotated[str, Depends(get_current_user_nuid)],
    current_user_id: Annotated[int, Depends(get_current_user_id)],
    db: Session = Depends(get_db),
):
    result = get_result_by_id(result_id, db)
    if result is None:
        raise HTTPException(status_code=404, detail="Result not found")
    # Verify session ownership
    _require_session_ownership(result.ExtractionSessionId, current_user_id, db)

    updated = update_result_review(
        result_id=result_id,
        reviewed_value=payload.reviewed_value,
        is_reviewed=payload.is_reviewed,
        is_incorrect=payload.is_incorrect,
        reviewer=current_user,
        db=db,
    )
    return updated


@router.post(
    "/results/{session_id}/approve-all",
    dependencies=[Depends(JWTBearer(_ALLOWED_ROLES))],
)
async def approve_all_high_confidence(
    session_id: int,
    threshold: float = Query(default=0.8, ge=0.0, le=1.0),
    current_user: Annotated[str, Depends(get_current_user_nuid)] = None,
    current_user_id: Annotated[int, Depends(get_current_user_id)] = None,
    db: Session = Depends(get_db),
):
    _require_session_ownership(session_id, current_user_id, db)
    count = bulk_approve_high_confidence(session_id, threshold, current_user, db)
    return {"approved_count": count}


@router.get(
    "/results/{session_id}/low-confidence-cases",
    dependencies=[Depends(JWTBearer(_ALLOWED_ROLES))],
)
async def get_low_confidence_cases(
    session_id: int,
    threshold: float = Query(default=0.8, ge=0.0, le=1.0),
    current_user_id: Annotated[int, Depends(get_current_user_id)] = None,
    db: Session = Depends(get_db),
):
    """Return distinct case IDs from the latest run that have any field below
    the confidence threshold or that have not yet been reviewed."""
    _require_session_ownership(session_id, current_user_id, db)
    case_ids = get_low_confidence_case_ids(session_id, threshold, db)
    return {"case_ids": case_ids, "count": len(case_ids)}


@router.get(
    "/results/{session_id}/incorrect-cases",
    dependencies=[Depends(JWTBearer(_ALLOWED_ROLES))],
)
async def get_incorrect_cases(
    session_id: int,
    current_user_id: Annotated[int, Depends(get_current_user_id)] = None,
    db: Session = Depends(get_db),
):
    """Return distinct case IDs from the latest run that have any field marked incorrect."""
    _require_session_ownership(session_id, current_user_id, db)
    case_ids = get_incorrect_case_ids(session_id, db)
    return {"case_ids": case_ids, "count": len(case_ids)}


# ──────────────────────────────────────────────────────────────────────────────
# Export
# ──────────────────────────────────────────────────────────────────────────────

@router.get(
    "/export/{session_id}",
    dependencies=[Depends(JWTBearer(_ALLOWED_ROLES))],
)
async def export_results(
    session_id: int,
    format: str = Query(default="csv", regex="^(csv|json|excel)$"),
    current_user_id: Annotated[int, Depends(get_current_user_id)] = None,
    db: Session = Depends(get_db),
):
    sess = _require_session_ownership(session_id, current_user_id, db)
    results = get_results_for_session(session_id, db)

    # Determine field order from schema
    field_order: List[str] = []
    if sess.SchemaJson:
        try:
            schema = json.loads(sess.SchemaJson)
            field_order = list(schema.keys())
        except Exception:
            pass

    # Build case_number → field → value table
    case_fields: Dict[str, Dict[str, Any]] = {}
    case_order: List[str] = []   # preserve first-seen order
    all_fields: set = set()
    for r in results:
        case_key = r.CaseNumber or str(r.CaseId)
        if case_key not in case_fields:
            case_fields[case_key] = {}
            case_order.append(case_key)
        value = r.ExtractedValue
        try:
            value = json.loads(value) if value is not None else None
        except Exception:
            pass
        # Flatten lists/dicts to a readable string
        if isinstance(value, (list, dict)):
            value = json.dumps(value)
        case_fields[case_key][r.FieldName] = value
        all_fields.add(r.FieldName)

    # Preserve schema field order, appending any extra fields alphabetically
    if field_order:
        extra = [f for f in sorted(all_fields) if f not in field_order]
        fields = field_order + extra
    else:
        fields = sorted(all_fields)

    if format == "json":
        rows = [
            {"case_number": cn, **case_fields[cn]}
            for cn in case_order
        ]
        return JSONResponse(content=rows)

    if format == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = (sess.Name or "Results")[:31]

        # Header row
        headers = ["Case Number"] + fields
        ws.append(headers)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for case_number in case_order:
            field_vals = case_fields[case_number]
            row = [case_number] + [field_vals.get(f) for f in fields]
            ws.append(row)

        # Auto-fit column widths (cap at 60)
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            width = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 60)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="extraction_{session_id}.xlsx"'
            },
        )

    # CSV
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=["case_number"] + fields, extrasaction="ignore")
    writer.writeheader()
    for case_number in case_order:
        row = {"case_number": case_number}
        for field in fields:
            row[field] = case_fields[case_number].get(field, "")
        writer.writerow(row)

    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={
            "Content-Disposition": f'attachment; filename="extraction_{session_id}.csv"'
        },
    )


# ──────────────────────────────────────────────────────────────────────────────
# AI field suggestion
# ──────────────────────────────────────────────────────────────────────────────

@router.post(
    "/suggest-fields",
    dependencies=[Depends(JWTBearer(_ALLOWED_ROLES))],
    response_model=List[FieldSuggestionVM],
)
async def suggest_fields(
    payload: FieldSuggestionRequest,
    current_user_id: Annotated[int, Depends(get_current_user_id)],
    current_role: Annotated[str, Depends(get_current_user_role)],
    db: Session = Depends(get_db),
):
    sample_text = payload.sample_text

    # If no sample text provided, grab first case from the session queue
    if not sample_text and payload.session_id:
        _require_session_ownership(payload.session_id, current_user_id, db)
        queue = get_queue(payload.session_id, db)
        if queue:
            labelled_text, _ = get_case_text_for_extraction(
                queue[0].CaseId, db, role=current_role
            )
            sample_text = labelled_text

    if not sample_text:
        raise HTTPException(
            status_code=400,
            detail="Provide either sample_text or a session_id with at least one case in the queue",
        )

    llm = get_llm_client()
    suggestions = await llm.suggest_fields(sample_text)
    return suggestions


# ──────────────────────────────────────────────────────────────────────────────
# Schema builder live preview (single-doc sync extraction)
# ──────────────────────────────────────────────────────────────────────────────

@router.post(
    "/preview",
    dependencies=[Depends(JWTBearer(_ALLOWED_ROLES))],
    response_model=ExtractionPreviewVM,
)
async def preview_extraction(
    payload: ExtractionPreviewRequest,
    current_user_id: Annotated[int, Depends(get_current_user_id)],
    current_role: Annotated[str, Depends(get_current_user_role)],
    db: Session = Depends(get_db),
):
    _require_session_ownership(payload.session_id, current_user_id, db)

    if not payload.extraction_schema:
        raise HTTPException(status_code=400, detail="Schema cannot be empty")

    labelled_text, _ = get_case_text_for_extraction(
        payload.case_id, db, role=current_role
    )

    if not labelled_text.strip():
        raise HTTPException(status_code=404, detail="No report text found for this case")

    llm = get_llm_client()
    extraction = await llm.extract(labelled_text, payload.extraction_schema)

    fields = {
        field_name: ExtractionPreviewFieldVM(
            value=fe.value,
            confidence=fe.confidence,
            provenance=fe.provenance,
        )
        for field_name, fe in extraction.items()
    }

    return ExtractionPreviewVM(
        case_id=payload.case_id,
        extracted_fields=fields,
        report_text=labelled_text,
    )


# ──────────────────────────────────────────────────────────────────────────────
# Case text
# ──────────────────────────────────────────────────────────────────────────────

@router.get(
    "/case/{case_id}/text",
    dependencies=[Depends(JWTBearer(_ALLOWED_ROLES))],
    response_model=CaseTextVM,
)
async def get_case_text(
    case_id: int,
    current_role: Annotated[str, Depends(get_current_user_role)],
    db: Session = Depends(get_db),
):
    labelled_text, segments = get_case_text_for_extraction(case_id, db, role=current_role)
    segment_vms = [
        CaseCommentSegmentVM(
            CaseCommentId=getattr(seg, "CaseCommentId", seg.Id),
            CommentType=seg.CommentType,
            CommentText=seg.CommentText,
        )
        for seg in segments
    ]
    return CaseTextVM(case_id=case_id, segments=segment_vms, full_text=labelled_text)
