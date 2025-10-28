# If on Python 2.X

from io import BytesIO
from typing import List

from exception.custom_exception import CustomException
from core.config import Settings
from openpyxl import Workbook, load_workbook
from viewmodel.cohort import CohortDetailsVM, CohortVMUpdate, CohortDataVM
from viewmodel.cohortPatient import CohortPatientVM
from viewmodel.cohortCase import CohortCaseVM
from core.constants import Constants


def read_excel(input: CohortVMUpdate):
    # save excel
    wb = Workbook()
    wb = load_workbook(filename=BytesIO(input.fileData))
    ws1 = wb["Data"]

    max_row = ws1.max_row
    cohortData: List[CohortDataVM] = []
    row_start = 4
    mrns = []

    type = ws1.cell(row=row_start - 1, column=1).value
    if input.type == Constants.CohortTypeMrn:
        if type != "MRN":
            raise CustomException(
                "Incorrect template used. Please use MRN template"
            )
    elif input.type == Constants.CohortTypeCase:
        if type != "Case":
            raise CustomException(
                "Incorrect template used. Please use Case template"
            )
    elif input.type == Constants.CohortTypeEpi:
        if type != "EID":
            raise CustomException(
                "Incorrect template used. Please use EID template"
            )

    # Will print a particular row value
    for i in range(max_row - 3):
        dataValue = ws1.cell(row=i + row_start, column=1).value
        if dataValue not in mrns:
            mrns.append(dataValue)

            item: CohortDataVM = CohortDataVM()
            item.data = dataValue
            item.cohortId = input.cohortId
            cohortData.append(item)
    return cohortData


def create_excel_patient_mrn(
    cohortId: int, cohort: CohortDetailsVM, data: List[CohortPatientVM]
):
    wb = Workbook()
    wb = load_workbook(
        f"{Settings.EXCEL_Template_DIRECTORY}{Settings.EXCEL_Cohort_MRN_Template_FILE}"
    )
    ws1 = wb["Data"]

    # Cohort details
    ws1.cell(1, 2, cohort.name)

    ws1.cell(1, 9, cohort.patientCount)
    ws1.cell(1, 11, cohort.matched)
    ws1.cell(1, 13, cohort.unmatched)

    row = 4
    for rowData in data:
        ws1.cell(row, 1, rowData.mrn)

        ws1.cell(row, 2, "Yes" if rowData.isfound else "No")
        row += 1

    file = f"CohortMRN_{cohortId}.xlsx"
    path = f"{Settings.EXCEL_Output_DIRECTORY}{file}"
    wb.save(path)
    return {"path": path, "file": file}


def create_excel_patient_epi(
    cohortId: int, cohort: CohortDetailsVM, data: List[CohortPatientVM]
):
    wb = Workbook()
    wb = load_workbook(
        f"{Settings.EXCEL_Template_DIRECTORY}{Settings.EXCEL_Cohort_EID_Template_FILE}"
    )
    ws1 = wb["Data"]

    # Cohort details
    ws1.cell(1, 2, cohort.name)

    ws1.cell(1, 9, cohort.patientCount)
    ws1.cell(1, 11, cohort.matched)
    ws1.cell(1, 13, cohort.unmatched)

    row = 4
    for rowData in data:
        ws1.cell(row, 1, rowData.epi)

        ws1.cell(row, 2, "Yes" if rowData.isfound else "No")
        row += 1

    file = f"CohortEID_{cohortId}.xlsx"
    path = f"{Settings.EXCEL_Output_DIRECTORY}{file}"
    wb.save(path)
    return {"path": path, "file": file}


def create_excel_case(
    cohortId: int, cohort: CohortDetailsVM, data: List[CohortCaseVM]
):
    wb = Workbook()
    wb = load_workbook(
        f"{Settings.EXCEL_Template_DIRECTORY}{Settings.EXCEL_Cohort_CASE_Template_FILE}"
    )
    ws1 = wb["Data"]

    # Cohort details
    ws1.cell(1, 2, cohort.name)

    ws1.cell(1, 9, cohort.caseCount)
    ws1.cell(1, 11, cohort.caseCountMatched)
    ws1.cell(1, 13, cohort.caseCount - cohort.caseCountMatched)

    row = 1

    # ws1.cell(
    #             3,
    #             1,
    #             "mrn"
    #         )

    # ws1.cell(
    #             3,
    #             2,
    #             "Found"
    #         )

    row = 4
    for rowData in data:
        ws1.cell(row, 1, rowData.case)

        ws1.cell(row, 2, "Yes" if rowData.isfound else "No")
        row += 1

    file = f"CohortCase_{cohortId}.xlsx"
    path = f"{Settings.EXCEL_Output_DIRECTORY}{file}"
    wb.save(path)
    return {"path": path, "file": file}


def create_mrn_template():
    return {
        "path": f"{Settings.EXCEL_Template_DIRECTORY}{Settings.EXCEL_Cohort_MRN_Template_FILE}",
        "file": "Cohort_MRN_Template.xlsx",
    }


def create_case_template():
    return {
        "path": f"{Settings.EXCEL_Template_DIRECTORY}{Settings.EXCEL_Cohort_CASE_Template_FILE}",
        "file": "Cohort_Case_Template.xlsx",
    }


def create_eid_template():
    return {
        "path": f"{Settings.EXCEL_Template_DIRECTORY}{Settings.EXCEL_Cohort_EID_Template_FILE}",
        "file": "Cohort_EID_Template.xlsx",
    }
