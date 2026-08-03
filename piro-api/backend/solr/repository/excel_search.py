# If on Python 2.X

import json
import os
from typing import Any, Dict, List, Optional
from urllib.parse import parse_qs, urlparse

from core.config import Settings
from db.repository.search import get_search
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE, Cell
from pytest import Session
from solr.models.document import document
from solr.repository.piro import read_result, search_Q
from viewmodel.solr.search import SearchFilterVM
from pysolr import Solr
from core.search_util import filter_str_object


def get_search_data(
    searchId: int,
    reasonCode: str,
    db: Session,
    solr: Solr,
    fields: str,
):
    search = get_search(searchId, db=db)
    advsearch = ""
    if search.AdvancedQuery is not None and search.AdvancedQuery != "":
        advsearch = filter_str_object(search.AdvancedQuery)
    query = search.SearchQuery
    parsed_url = urlparse(query)
    parsed_q = parse_qs(parsed_url.query)
    json_object = json.loads(parsed_q["searchFilter"][0])
    filters = []
    for item in json_object:
        # obj = SearchFilterVM
        # obj.category = obj['category']
        # obj.field = obj['field']
        # obj.search = obj['search']
        # obj.andcondition = obj['andcondition']
        obj = SearchFilterVM.parse_obj(item)

        filters.append(obj)

    # Add deceased filter
    if reasonCode == "DEC":
        filters.append(
            SearchFilterVM(
                field="isdeceased",
                search="Deceased",
                category="isdeceased",
                andcondition=True,
                displaysingular="",
            )
        )

    docs = search_Q(
        input_arr=filters,
        input_adv=advsearch,
        mrn=search.MRN,
        sortBy="",
        sortOrder="",
        page=1,
        count=Settings.EXCEL_Output_Records,
        db=db,
        solr=solr,
        finalRtf=False,
        fields=fields,
    )
    return list(docs["items"])


def get_case_data_by_ids(
    case_ids: List[int],
    db: Session,
    solr: Solr,
    fields: str,
):
    """Fetch Solr document data for an explicit list of case ids.

    Used for LLM-assisted data requests, where the source of cases is an
    ExtractionSession's queue rather than a Saved Search query.
    """
    if not case_ids:
        return []

    ids_query = " OR ".join(str(c) for c in case_ids)
    query = f"caseid:({ids_query})"
    solr_query_params = {"start": 0, "rows": len(case_ids)}
    results = solr.search(query, fl=fields, **solr_query_params)

    docs: List[document] = []
    for result in results:
        docs.append(read_result(result, db=db))
    return docs


def create_excel(
    searchId: int,
    data: List[document],
    fields: [],
    extraction_fields: Optional[List[str]] = None,
    extraction_data: Optional[Dict[str, Dict[str, Any]]] = None,
):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Data"

    offset_row = 0
    offset_col = 0
    col = 1
    row = 1
    for field in fields:
        ws1.cell(row, col + offset_col, field.DataFieldDisplayName)
        col = col + 1

    extraction_fields = extraction_fields or []
    extraction_data = extraction_data or {}
    extraction_col_start = col
    for extraction_field in extraction_fields:
        ws1.cell(row, col + offset_col, extraction_field)
        col = col + 1

    row = 2
    for rowData in data:
        col = 1
        rowDic = rowData.__dict__
        for field in fields:
            if field.DataFieldSolrField in rowDic:
                if rowDic[field.DataFieldSolrField] is None:
                    ws1.cell(
                        row + offset_row,
                        col + offset_col,
                        ILLEGAL_CHARACTERS_RE.sub(r"", ""),
                    )
                else:
                    # delimiter1 = "   ||||   "
                    # delimiter2 = "||--||"
                    str_data = ILLEGAL_CHARACTERS_RE.sub(
                        r"", str(rowDic[field.DataFieldSolrField])
                    )
                    str_data = str_data.replace("   ||||   ", "\n\n")
                    str_data = str_data.replace("||--||", "\n\n")
                    if field.DataFieldSolrField == "final":
                        str_data = str_data.replace("-", "\n-")
                    cell: Cell = ws1.cell(
                        row + offset_row,
                        col + offset_col,
                        str_data,
                    )
                    # Set the cell's data type to String ('s') to avoid text
                    # being interpreted as a formula.
                    cell.data_type = "s"
            else:
                ws1.cell(
                    row + offset_row,
                    col + offset_col,
                    ILLEGAL_CHARACTERS_RE.sub(r"", ""),
                )
            col = col + 1

        if extraction_fields:
            case_key = str(rowDic.get("caseid", ""))
            row_extracted = extraction_data.get(case_key, {})
            col = extraction_col_start
            for extraction_field in extraction_fields:
                value = row_extracted.get(extraction_field)
                str_data = (
                    ""
                    if value is None
                    else ILLEGAL_CHARACTERS_RE.sub(r"", str(value))
                )
                cell: Cell = ws1.cell(row + offset_row, col + offset_col, str_data)
                cell.data_type = "s"
                col = col + 1

        row += 1

    file = f"PIRO_{searchId}.xlsx"
    output_dir = Settings.EXCEL_Output_DIRECTORY or ""
    if output_dir != "":
        os.makedirs(output_dir, exist_ok=True)
    path = os.path.join(output_dir, file)
    wb.save(path)
    return {"path": path, "file": file}
